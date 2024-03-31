import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from openpyxl import load_workbook
import math

wb = load_workbook(filename="uloha3.xlsx", data_only=True)
ws_vystupni = wb["Vystupni"]
ws_proudova = wb["Proudova"]
ws_vstupni = wb["Vstupni"]

# Uce je společné pro všechny výstupní charakteristiky
Uce = np.empty(20)

i = 0
for row in ws_vystupni.iter_rows(min_row=4, max_row=23, min_col=2, max_col=2, values_only=True):
    for value in row:
        Uce[i] = value
        i += 1

print(Uce)

# Ib = 20 uA
Ic_20 = np.empty(20)

i = 0
for row in ws_vystupni.iter_rows(min_row=4, max_row=23, min_col=3, max_col=3, values_only=True):
    for value in row:
        Ic_20[i] = value
        i += 1

print(Ic_20)

# Ib = 30 uA
Ic_30 = np.empty(20)

i = 0
for row in ws_vystupni.iter_rows(min_row=4, max_row=23, min_col=6, max_col=6, values_only=True):
    for value in row:
        Ic_30[i] = value
        i += 1

print(Ic_30)

# Ib = 40 uA
Ic_40 = np.empty(20)

i = 0
for row in ws_vystupni.iter_rows(min_row=4, max_row=23, min_col=9, max_col=9, values_only=True):
    for value in row:
        Ic_40[i] = value
        i += 1

print(Ic_40)

# Ib = 50 uA
Ic_50 = np.empty(20)

i = 0
for row in ws_vystupni.iter_rows(min_row=4, max_row=23, min_col=12, max_col=12, values_only=True):
    for value in row:
        Ic_50[i] = value
        i += 1

print(Ic_50)

# Proudova charakteristika
Ib_proudova = np.empty(8)
Ic_proudova = np.empty(8)

i = 0
for row in ws_proudova.iter_rows(min_row=4, max_row=11, min_col=2, max_col=2, values_only=True):
    for value in row:
        Ib_proudova[i] = value
        i += 1
i = 0
for row in ws_proudova.iter_rows(min_row=4, max_row=11, min_col=3, max_col=3, values_only=True):
    for value in row:
        Ic_proudova[i] = value
        i += 1

print(Ib_proudova, Ic_proudova)

# Vstupni charakteristika
Ib_vstupni = np.empty(9)
Ube_vstupni = np.empty(9)

i = 0
for row in ws_vstupni.iter_rows(min_row=4, max_row=12, min_col=2, max_col=2, values_only=True):
    for value in row:
        Ib_vstupni[i] = value
        i += 1
i = 0
for row in ws_vstupni.iter_rows(min_row=4, max_row=12, min_col=3, max_col=3, values_only=True):
    for value in row:
        Ube_vstupni[i] = value
        i += 1

print(Ib_vstupni, Ube_vstupni)

# Plot drawing
fig = plt.figure()
grid = gridspec.GridSpec(nrows=2, ncols=2, hspace=0, wspace=0, figure=fig)

i_quadrant = fig.add_subplot(grid[0, 1], zorder=3)
i_quadrant.plot(Uce, Ic_20, color='red', marker='x', markersize='5', markeredgecolor='darkred', label="$I_B = 20 \mu A$")
i_quadrant.plot(Uce, Ic_30, color='limegreen', marker='x', markersize='5', markeredgecolor='darkgreen', label="$I_B = 30 \mu A$")
i_quadrant.plot(Uce, Ic_40, color='blue', marker='x', markersize='5', markeredgecolor='midnightblue', label="$I_B = 40 \mu A$")
i_quadrant.plot(Uce, Ic_50, color='mediumorchid', marker='x', markersize='5', markeredgecolor='indigo', label="$I_B = 50 \mu A$")
i_quadrant.text(7, Ic_50[-4], "$I_B = 50 \mu A$", rotation=math.degrees(math.atan(Ic_50[-3] - Ic_50[-4])), transform_rotates_text=True, color="mediumorchid", horizontalalignment="left", verticalalignment="bottom")
i_quadrant.text(7, Ic_40[-4], "$I_B = 40 \mu A$", rotation=math.degrees(math.atan(Ic_40[-3] - Ic_40[-4])), transform_rotates_text=True, color="blue", horizontalalignment="left", verticalalignment="bottom")
i_quadrant.text(7, Ic_30[-4], "$I_B = 30 \mu A$", rotation=math.degrees(math.atan(Ic_30[-3] - Ic_30[-4])), transform_rotates_text=True, color="limegreen", horizontalalignment="left", verticalalignment="bottom")
i_quadrant.text(7, Ic_20[-4], "$I_B = 20 \mu A$", rotation=math.degrees(math.atan(Ic_20[-3] - Ic_20[-4])), transform_rotates_text=True, color="red", horizontalalignment="left", verticalalignment="bottom")
i_quadrant.margins(0)

ii_quadrant = fig.add_subplot(grid[0, 0], zorder=2, sharey=i_quadrant)
ii_quadrant.plot(Ib_proudova, Ic_proudova, color='red', marker='x', markersize='5', markeredgecolor='darkred')
ii_quadrant.plot([20, 30, 40, 50], [5.92, 8.883, 11.919, 14.925], color='blue', marker='x', markersize='5', markeredgecolor='midnightblue')

ii_quadrant.margins(0)

iii_quadrant = fig.add_subplot(grid[1, 0], zorder=0, sharex=ii_quadrant)
iii_quadrant.plot(Ib_vstupni, Ube_vstupni, color='red', marker='x', markersize='5', markeredgecolor='darkred')
iii_quadrant.margins(0)

iv_quadrant = fig.add_subplot(grid[1, 1], zorder=1, sharex=i_quadrant, sharey=iii_quadrant)
#iv_quadrant.plot(x, y)
iv_quadrant.margins(0)

# Axes box settings
i_quadrant.spines['top'].set_visible(False)
i_quadrant.spines['right'].set_visible(False)
ii_quadrant.spines['top'].set_visible(False)
ii_quadrant.spines['left'].set_visible(False)
iii_quadrant.spines['left'].set_visible(False)
iii_quadrant.spines['bottom'].set_visible(False)
iv_quadrant.spines['right'].set_visible(False)
iv_quadrant.spines['bottom'].set_visible(False)

# Tick settings
ii_quadrant.tick_params(axis='both', labelleft=False)
ii_quadrant.tick_params(axis='y', left=False)
ii_quadrant.tick_params(axis='x', direction='in', pad=-15)

iii_quadrant.tick_params(axis='both', labelleft=False, labelbottom=False, left=False, bottom=False)

iv_quadrant.tick_params(axis='both', labelbottom=False, bottom=False)
iv_quadrant.tick_params(axis='y', direction='in', pad=-22)

plt.setp(i_quadrant.get_xticklabels()[0], visible=False)
plt.setp(i_quadrant.get_yticklabels()[0], visible=False)
#plt.setp(ii_quadrant.get_xticklabels()[0], visible=False)
plt.setp(iv_quadrant.get_yticklabels()[0], visible=False)

i_quadrant.set_xticks(np.arange(0, 12, 2))
i_quadrant.set_yticks(np.arange(0, 25, 5))
ii_quadrant.set_xticks(np.arange(20, 80, 10))
iv_quadrant.set_yticks(np.arange(0, 1, 0.2))

# Axis settings
i_quadrant.set_xlim([0, 11])
i_quadrant.set_ylim([0, 25])
ii_quadrant.set_xlim([0, 80])
ii_quadrant.set_ylim([0, 25])
iii_quadrant.set_xlim([0, 80])
iii_quadrant.set_ylim([0, 0.9])

i_quadrant.set_xlabel("$U_{CE}\ (V)$", loc='right')
i_quadrant.set_ylabel("$I_C\ (mA)$", loc='top')
ii_quadrant.set_xlabel("$I_B\ (\mu A)$", loc='left')
iv_quadrant.set_ylabel("$U_{BE}\ (V)$", loc='bottom')

i_quadrant.set_title("Výstupní", y=1, pad=-14)
ii_quadrant.set_title("Proudová\npřevodní", y=1, pad=-28)
iii_quadrant.set_title("Vstupní", y=0)

iii_quadrant.invert_xaxis()
iii_quadrant.invert_yaxis()

plt.show()