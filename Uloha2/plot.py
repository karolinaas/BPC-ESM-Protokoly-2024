from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt

wb = load_workbook(filename="uloha2.xlsx", data_only=True)
ws = wb["List1"]

Adb = np.empty(22)
Ur = np.empty(22)
i = 0

# Reading excel
for row in ws.iter_rows(min_row=3, max_row=24, min_col=2, max_col=2, values_only=True):
    for value in row:
        Adb[i] = value
        i += 1
i = 0
for row in ws.iter_rows(min_row=3, max_row=24, min_col=3, max_col=3, values_only=True):
    for value in row:
        Ur[i] = value
        i += 1

print(Adb)
print(Ur)

plt.plot(np.sort(Ur), np.sort(Adb), color='red', marker='x', markersize='6', markeredgecolor='darkred')
plt.ylabel("$A_{dB}\ (dB)$")
plt.xlabel("$U_R\ (V)$")
plt.title("Přenosová charakteristika diody 1N4148")
plt.grid(True)
plt.show()